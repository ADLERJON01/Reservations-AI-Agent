"""Stratified sampler for the gold-label set (evaluation track).

Draws a stratified sample from the curated 77-email pool
(`outputs/manual_labels_template.csv`, v1.0.0 taxonomy, `my_*` columns empty) and
writes a focused labeling sheet for hand-labeling (category + 5 facets).

Why stratified: the rare categories (payment ~4%, inventory ~3%, system ~4%,
other ~1%) would be near-absent in a representative draw, so we keep ALL of them
and trim the large buckets — then per-category metrics are meaningful and the
scorer reweights to the true distribution for a headline number.

Deterministic (fixed seed). Run:
    python -m app.eval.sample_gold              # ~50, stratified
    python -m app.eval.sample_gold --target all # label the full 77-pool
"""
from __future__ import annotations

import argparse
import csv
import random
from collections import defaultdict
from pathlib import Path
from typing import get_args

from app.models.llm_output import (Category, ExpectsHuman, LifecycleStage,
                                    RequestType, SenderType, Urgency)

OUTPUTS_DIR = Path(__file__).resolve().parents[2] / "outputs"
POOL_CSV = OUTPUTS_DIR / "manual_labels_template.csv"
OUT_DIR = OUTPUTS_DIR / "gold"

SEED = 7
TARGET_TOTAL = 50

# Rare categories: keep ALL of them regardless of target.
KEEP_ALL = {
    "payment_billing_or_rate_issue",
    "inventory_availability_or_stop_sales",
    "system_or_channel_delivery_exception",
    "other_or_unclear",
}
# Large buckets: trim to fill the remaining budget, proportional to pool size.
TRIM = [
    "service_or_information_inquiry",
    "booking_notification",
    "booking_change_or_cancellation",
]

# Documented fork-point emails (borderline_cases.md) — force-include any in the pool.
BORDERLINES = {f"email_{n}" for n in (
    98, 335, 336, 345, 369, 99, 155, 161, 24, 337, 338, 376, 378, 377,
    326, 327, 324, 325, 4, 127, 133, 72, 368, 358, 157, 204, 299, 300, 361)}

# The sheet is a clean one-row-per-email grid (NO multi-line body — that breaks
# Excel). The full bodies live in the companion gold_emails.md for reading.
GOLD_COLUMNS = [
    "email_id", "subject_full", "body_preview",
    "proposed_category", "proposal_method",          # heuristic HINTS (not the LLM)
    "gold_category", "gold_sender_type", "gold_request_type",
    "gold_booking_lifecycle_stage", "gold_expects_human_response", "gold_urgency_signal",
    "ambiguous", "notes",
]
PREVIEW_CHARS = 180


def _oneline(text: str, limit: int | None = None) -> str:
    """Collapse all whitespace/newlines to single spaces (safe for one CSV cell)."""
    flat = " ".join((text or "").split())
    return flat[:limit] + "…" if (limit and len(flat) > limit) else flat


# Dropdown options pulled from the LOCKED taxonomy enums (single source of truth —
# can't drift from the schema). Each maps a gold_* column to its allowed values.
GOLD_FACETS = [
    ("gold_category", list(get_args(Category))),
    ("gold_sender_type", list(get_args(SenderType))),
    ("gold_request_type", list(get_args(RequestType))),
    ("gold_booking_lifecycle_stage", list(get_args(LifecycleStage))),
    ("gold_expects_human_response", list(get_args(ExpectsHuman))),
    ("gold_urgency_signal", list(get_args(Urgency))),
    ("ambiguous", ["Y", "N"]),
]


def _load_pool() -> list[dict]:
    with open(POOL_CSV, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def select(rows: list[dict], target) -> list[dict]:
    by_id = {r["email_id"]: r for r in rows}
    by_cat: dict[str, list[dict]] = defaultdict(list)
    for r in sorted(rows, key=lambda r: r["email_id"]):
        by_cat[r["proposed_category"]].append(r)

    if target == "all":
        return sorted(rows, key=lambda r: r["email_id"])

    rng = random.Random(SEED)
    chosen: dict[str, dict] = {}

    # 1) force-include borderlines present in the pool
    for eid in sorted(BORDERLINES):
        if eid in by_id:
            chosen[eid] = by_id[eid]
    # 2) keep ALL rare-category rows
    for cat in KEEP_ALL:
        for r in by_cat.get(cat, []):
            chosen[r["email_id"]] = r
    # 3) fill the remaining budget from the large buckets, proportional to pool size
    remaining = max(0, target - len(chosen))
    pool_trim = {c: [r for r in by_cat.get(c, []) if r["email_id"] not in chosen] for c in TRIM}
    total_trim = sum(len(v) for v in pool_trim.values()) or 1
    for cat in TRIM:
        take = round(remaining * len(pool_trim[cat]) / total_trim)
        for r in rng.sample(pool_trim[cat], min(take, len(pool_trim[cat]))):
            chosen[r["email_id"]] = r
    return sorted(chosen.values(), key=lambda r: r["email_id"])


def write_sheet(selected: list[dict]) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / "gold_labeling_sheet.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=GOLD_COLUMNS, extrasaction="ignore")
        w.writeheader()
        for r in selected:
            row = {k: _oneline(r.get(k, "")) for k in ("email_id", "subject_full",
                                                       "proposed_category", "proposal_method")}
            row["body_preview"] = _oneline(r.get("body_clean", ""), PREVIEW_CHARS)
            for g in ("gold_category", "gold_sender_type", "gold_request_type",
                      "gold_booking_lifecycle_stage", "gold_expects_human_response",
                      "gold_urgency_signal", "ambiguous", "notes"):
                row[g] = ""                     # blank → hand-labeled
            w.writerow(row)
    return out


def write_reading_doc(selected: list[dict]) -> Path:
    """Full email bodies for reading while labeling (one section per email)."""
    out = OUT_DIR / "gold_emails.md"
    lines = ["# Gold-set emails — full bodies (read while labeling)\n"]
    for r in selected:
        lines += [
            f"\n## {r['email_id']}",
            f"**Subject:** {_oneline(r.get('subject_full',''))}",
            f"**Heuristic hint (not the answer):** {r.get('proposed_category','')}\n",
            "```",
            (r.get("body_clean", "") or "").strip(),
            "```",
        ]
    out.write_text("\n".join(lines), encoding="utf-8")
    return out


def write_xlsx(selected: list[dict]) -> Path | None:
    """Excel sheet with a dropdown (data-validation list) on each gold_* column.

    Allowed values live on a hidden 'Lists' sheet and are referenced by range —
    request_type's 10 values exceed Excel's 255-char inline-list limit."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("  (.xlsx skipped — `pip install openpyxl`)")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "labels"
    ws.append(GOLD_COLUMNS)
    for r in selected:
        ws.append([
            _oneline(r.get("email_id", "")), _oneline(r.get("subject_full", "")),
            _oneline(r.get("body_clean", ""), PREVIEW_CHARS),
            _oneline(r.get("proposed_category", "")), _oneline(r.get("proposal_method", "")),
            "", "", "", "", "", "", "", "",                  # gold_* + ambiguous + notes
        ])
    last = len(selected) + 1                                  # header is row 1

    for c in range(1, len(GOLD_COLUMNS) + 1):
        ws.cell(1, c).font = Font(bold=True)
    ws.freeze_panes = "A2"
    widths = {"email_id": 12, "subject_full": 42, "body_preview": 52, "proposed_category": 26,
              "proposal_method": 14, "gold_category": 30, "gold_sender_type": 24,
              "gold_request_type": 30, "gold_booking_lifecycle_stage": 26,
              "gold_expects_human_response": 24, "gold_urgency_signal": 22,
              "ambiguous": 10, "notes": 30}
    for i, col in enumerate(GOLD_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 16)
    wrap = Alignment(wrap_text=True, vertical="top")
    for col in ("subject_full", "body_preview", "notes"):
        ci = GOLD_COLUMNS.index(col) + 1
        for row in range(2, last + 1):
            ws.cell(row, ci).alignment = wrap

    lists = wb.create_sheet("Lists")
    lists.sheet_state = "hidden"
    for j, (name, values) in enumerate(GOLD_FACETS, start=1):
        lcol = get_column_letter(j)
        lists.cell(1, j, name)
        for k, v in enumerate(values, start=2):
            lists.cell(k, j, v)
        dv = DataValidation(type="list", allow_blank=True, showErrorMessage=True,
                            formula1=f"Lists!${lcol}$2:${lcol}${1 + len(values)}")
        dv.errorTitle, dv.error = "Invalid label", "Pick a value from the dropdown."
        ws.add_data_validation(dv)
        gcol = get_column_letter(GOLD_COLUMNS.index(name) + 1)
        dv.add(f"{gcol}2:{gcol}{last}")

    out = OUT_DIR / "gold_labeling_sheet.xlsx"
    wb.save(out)
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--target", default=str(TARGET_TOTAL),
                    help="number of emails, or 'all' for the full 77-pool")
    args = ap.parse_args()
    target = "all" if args.target == "all" else int(args.target)

    rows = _load_pool()
    selected = select(rows, target)
    sheet = write_sheet(selected)
    reading = write_reading_doc(selected)
    xlsx = write_xlsx(selected)

    counts: dict[str, int] = defaultdict(int)
    borderline_hit = 0
    for r in selected:
        counts[r["proposed_category"]] += 1
        if r["email_id"] in BORDERLINES:
            borderline_hit += 1
    print(f"pool={len(rows)}  selected={len(selected)}  "
          f"borderlines_included={borderline_hit}")
    print(f"→ sheet:   {sheet}")
    if xlsx:
        print(f"→ xlsx:    {xlsx}  (dropdowns on gold_* columns)")
    print(f"→ reading: {reading}")
    for cat, n in sorted(counts.items(), key=lambda kv: -kv[1]):
        print(f"  {n:3d}  {cat}")


if __name__ == "__main__":
    main()
