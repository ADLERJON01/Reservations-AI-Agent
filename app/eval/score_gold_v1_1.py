"""Score v1.1 predictions vs the gold (OFFLINE).

The gold lives in the v2 sheet (gold_labeling_sheet_v2.xlsx). v1.1 has only 7
categories, so the gold category is COLLAPSED deterministically:
  {knowledge_policy_inquiry, sales_availability_or_quote_inquiry,
   guest_service_or_ancillary_request, thread_closure_or_acknowledgment}
   -> service_or_information_inquiry
(plus the confirmed email_299/email_300 -> booking_notification fix). All other facets
are read as-is. RAG gate = service_or_information_inquiry AND kb_policy.

    python -m app.eval.score_gold_v1_1
"""
from __future__ import annotations

import csv
import warnings
from pathlib import Path

from openpyxl import load_workbook

OUT_DIR = Path(__file__).resolve().parents[2] / "outputs" / "gold"
SHEET = OUT_DIR / "gold_labeling_sheet_v2.xlsx"
PRED = OUT_DIR / "gold_predictions_v1_1.csv"

V1_CATEGORY_BASELINE = 0.796     # v1 enriched original (outputs/gold/gold_metrics.md)
V21_CATEGORY = 0.694             # v2.1 native 10-cat (gold_metrics_v2.1.md)

# v2 (10-cat) -> v1 (7-cat) collapse.
COLLAPSE = {
    "knowledge_policy_inquiry": "service_or_information_inquiry",
    "sales_availability_or_quote_inquiry": "service_or_information_inquiry",
    "guest_service_or_ancillary_request": "service_or_information_inquiry",
    "thread_closure_or_acknowledgment": "service_or_information_inquiry",
}
FIX = {"email_299": "booking_notification", "email_300": "booking_notification"}

FACETS = [
    ("category", "gold_category", "pred_category"),       # gold collapsed below
    ("request_type", "gold_request_type", "pred_request_type"),
    ("inquiry_answer_source", "gold_inquiry_answer_source", "pred_inquiry_answer_source"),
    ("lifecycle", "gold_booking_lifecycle_stage", "pred_booking_lifecycle_stage"),
    ("requires_human_followup", "gold_requires_human_followup", "pred_requires_human_followup"),
    ("sender_type", "gold_sender_type", "pred_sender_type"),
    ("urgency_signal", "gold_urgency_signal", "pred_urgency_signal"),
]


def _collapse(cat: str) -> str:
    return COLLAPSE.get(cat, cat)


def _rag(cat: str, src: str) -> bool:
    return cat == "service_or_information_inquiry" and src == "kb_policy"


def load_gold() -> dict[str, dict]:
    warnings.filterwarnings("ignore")
    ws = load_workbook(SHEET)["labels"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    gold = {}
    for r in rows[1:]:
        eid = r[idx["email_id"]]
        if not eid:
            continue
        g = {k: (r[idx[k]] or "") for k in idx}
        g["gold_category"] = _collapse(FIX.get(eid, r[idx["gold_category"]] or ""))
        g["_ambiguous"] = str(r[idx["ambiguous"]] or "").strip().upper() == "Y"
        gold[eid] = g
    return gold


def per_class(pairs):
    classes = sorted(set(g for g, _ in pairs) | set(p for _, p in pairs))
    out = {}
    for c in classes:
        tp = sum(1 for g, p in pairs if g == c and p == c)
        fp = sum(1 for g, p in pairs if g != c and p == c)
        fn = sum(1 for g, p in pairs if g == c and p != c)
        support = sum(1 for g, _ in pairs if g == c)
        prec = tp / (tp + fp) if tp + fp else 0.0
        rec = tp / (tp + fn) if tp + fn else 0.0
        f1 = 2 * prec * rec / (prec + rec) if prec + rec else 0.0
        out[c] = (prec, rec, f1, support)
    return out


def main() -> None:
    if not PRED.exists():
        raise SystemExit(f"No predictions at {PRED}. Run: python -m app.eval.predict_gold_v1_1")
    gold = load_gold()
    pred = {row["email_id"]: row for row in csv.DictReader(open(PRED, encoding="utf-8"))}
    ids = [e for e in gold if e in pred]

    lines = ["# Gold-set metrics — taxonomy_v1_1 (v1 7-cat + inquiry_answer_source patch)\n",
             f"Scored **{len(ids)}** emails. Gold category collapsed v2->v1 (7-cat).\n"]

    # facet accuracy
    lines += ["## Facet accuracy\n", "| facet | accuracy | correct/total |", "|---|---|---|"]
    facc = {}
    for name, gcol, pcol in FACETS:
        correct = sum(1 for e in ids if gold[e][gcol] == pred[e][pcol])
        facc[name] = correct / len(ids)
        lines.append(f"| {name} | {facc[name]:.1%} | {correct}/{len(ids)} |")
    lines.append("")

    cat_acc = facc["category"]
    unamb = [e for e in ids if not gold[e]["_ambiguous"]]
    cat_un = sum(1 for e in unamb if gold[e]["gold_category"] == pred[e]["pred_category"]) / len(unamb)
    lines.append(f"**Category {cat_acc:.1%} (all) | {cat_un:.1%} (unambiguous).** "
                 f"vs v1 enriched {V1_CATEGORY_BASELINE:.1%}, vs v2.1 {V21_CATEGORY:.1%}.\n")

    # per-category P/R/F1
    pairs = [(gold[e]["gold_category"], pred[e]["pred_category"] or "(invalid)") for e in ids]
    pc = per_class(pairs)
    lines += ["## Category precision / recall / F1\n", "| category | P | R | F1 | support |", "|---|---|---|---|---|"]
    for c, (p, r, f1, s) in sorted(pc.items(), key=lambda kv: -kv[1][3]):
        lines.append(f"| {c} | {p:.2f} | {r:.2f} | {f1:.2f} | {s} |")
    macro = sum(v[2] for v in pc.values()) / len(pc) if pc else 0.0
    lines.append(f"\n**Macro-F1: {macro:.2f}**\n")

    # RAG gate
    gr = {e: _rag(gold[e]["gold_category"], gold[e]["gold_inquiry_answer_source"]) for e in ids}
    prr = {e: _rag(pred[e]["pred_category"], pred[e]["pred_inquiry_answer_source"]) for e in ids}
    tp = [e for e in ids if gr[e] and prr[e]]
    fp = [e for e in ids if not gr[e] and prr[e]]
    fn = [e for e in ids if gr[e] and not prr[e]]
    prec = len(tp) / (len(tp) + len(fp)) if (tp or fp) else 0.0
    rec = len(tp) / (len(tp) + len(fn)) if (tp or fn) else 0.0
    lines += ["## RAG-candidate gate (service_or_information_inquiry AND kb_policy)\n",
              f"- gold candidates: **{sum(gr.values())}** | predicted: **{sum(prr.values())}**",
              f"- **precision {prec:.0%}** (TP {len(tp)}/{len(tp)+len(fp)}) | **recall {rec:.0%}** "
              f"(TP {len(tp)}/{len(tp)+len(fn)})",
              f"- FALSE candidates (dangerous) [{len(fp)}]: {fp or 'none'}",
              f"- missed candidates [{len(fn)}]: {fn or 'none'}\n"]

    # misclassifications
    errs = [(e, gold[e]["gold_category"], pred[e]["pred_category"] or "(invalid)")
            for e in ids if gold[e]["gold_category"] != (pred[e]["pred_category"] or "(invalid)")]
    lines.append(f"## Category misclassifications ({len(errs)})\n")
    for e, g, p in errs:
        flag = " [ambiguous]" if gold[e]["_ambiguous"] else ""
        lines.append(f"- `{e}`: gold **{g}** → pred **{p}**{flag}")
    lines.append("")

    out = OUT_DIR / "gold_metrics_v1_1.md"
    out.write_text("\n".join(lines), encoding="utf-8")

    print(f"scored {len(ids)} emails (v1.1)")
    print(f"category: {cat_acc:.1%} (all) | {cat_un:.1%} (unamb) | macro-F1 {macro:.2f}"
          f"  [v1 {V1_CATEGORY_BASELINE:.0%}, v2.1 {V21_CATEGORY:.0%}]")
    print(f"RAG gate: precision {prec:.0%} | recall {rec:.0%} | false {len(fp)} | missed {len(fn)}")
    for name in ("inquiry_answer_source", "request_type", "requires_human_followup",
                 "lifecycle", "sender_type", "urgency_signal"):
        print(f"  {name:24s} {facc[name]:.1%}")
    print(f"→ {out}")


if __name__ == "__main__":
    main()