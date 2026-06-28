"""Build the v2.0.0 gold re-labeling sheet — SAME 49 emails as v1, new taxonomy.

Reads the v1 labeled sheet (outputs/gold/gold_labeling_sheet.xlsx) and emits a v2
sheet with dropdowns on every gold_* column, so the v2 run is scored on the identical
email set (apples-to-apples vs the v1 baseline, same model).

Work-minimising design:
- CARRY OVER unchanged labels as defaults: sender_type, urgency_signal.
- CARRY OVER prior judgments as editable starting points: request_type,
  booking_lifecycle_stage (review under the new "stage of the referenced booking"
  convention), requires_human_followup (was expects_human_response).
- SHOW v1 category as a read-only reference column (guides the 4-way split of the old
  service bucket).
- BLANK the two genuinely new fields to relabel fresh: gold_category (10 values) and
  gold_inquiry_answer_source (5 values). ambiguous is re-assessed (blank); notes carry.

Read the full bodies in the existing outputs/gold/gold_emails.md (same 49 emails).

    python -m app.eval.sample_gold_v2
"""
from __future__ import annotations

import warnings
from pathlib import Path
from typing import get_args

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.models.llm_output_v2 import (Category, InquiryAnswerSource, LifecycleStage,
                                       RequestType, RequiresHumanFollowup, SenderType, Urgency)

OUT_DIR = Path(__file__).resolve().parents[2] / "outputs" / "gold"
V1_SHEET = OUT_DIR / "gold_labeling_sheet.xlsx"
V2_SHEET = OUT_DIR / "gold_labeling_sheet_v2.xlsx"

# v2 column order. To-fill-fresh columns are placed first (after context) so the eye
# lands on them.
COLUMNS = [
    "email_id", "subject_full", "body_preview",
    "proposed_category", "proposal_method",       # heuristic hints (from v1)
    "v1_gold_category",                            # READ-ONLY reference (guides the split)
    "gold_category",                              # BLANK — relabel (10 values)
    "gold_inquiry_answer_source",                # BLANK — new (5 values)
    "gold_request_type",                          # carried (review)
    "gold_booking_lifecycle_stage",               # carried (review: new convention)
    "gold_sender_type",                           # carried (unchanged)
    "gold_requires_human_followup",               # carried from expects_human_response
    "gold_urgency_signal",                         # carried (unchanged)
    "ambiguous",                                   # BLANK — reassess
    "notes",                                       # carried (reference)
]

# gold_* column -> allowed dropdown values (from the v2 enums; single source of truth).
FACETS = [
    ("gold_category", list(get_args(Category))),
    ("gold_inquiry_answer_source", list(get_args(InquiryAnswerSource))),
    ("gold_request_type", list(get_args(RequestType))),
    ("gold_booking_lifecycle_stage", list(get_args(LifecycleStage))),
    ("gold_sender_type", list(get_args(SenderType))),
    ("gold_requires_human_followup", list(get_args(RequiresHumanFollowup))),
    ("gold_urgency_signal", list(get_args(Urgency))),
    ("ambiguous", ["Y", "N"]),
]

BLANK = PatternFill("solid", fgColor="FFF2CC")   # highlight the to-fill-fresh columns


def _read_v1() -> list[dict]:
    warnings.filterwarnings("ignore")             # openpyxl drops the DV extension on read
    ws = load_workbook(V1_SHEET)["labels"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    out = []
    for r in rows[1:]:
        if not r[idx["email_id"]]:
            continue
        out.append({h: (r[i] if r[i] is not None else "") for h, i in idx.items()})
    return out


def main() -> None:
    v1 = _read_v1()
    wb = Workbook()
    ws = wb.active
    ws.title = "labels"
    ws.append(COLUMNS)

    for r in v1:
        ws.append([
            r.get("email_id", ""), r.get("subject_full", ""), r.get("body_preview", ""),
            r.get("proposed_category", ""), r.get("proposal_method", ""),
            r.get("gold_category", ""),                       # v1_gold_category (reference)
            "",                                               # gold_category (BLANK)
            "",                                               # gold_inquiry_answer_source (BLANK)
            r.get("gold_request_type", ""),                   # carried
            r.get("gold_booking_lifecycle_stage", ""),        # carried
            r.get("gold_sender_type", ""),                    # carried
            r.get("gold_expects_human_response", ""),         # carried -> requires_human_followup
            r.get("gold_urgency_signal", ""),                 # carried
            "",                                               # ambiguous (reassess)
            r.get("notes", ""),                               # carried
        ])
    last = len(v1) + 1

    # header + widths + wrapping
    for c in range(1, len(COLUMNS) + 1):
        ws.cell(1, c).font = Font(bold=True)
    ws.freeze_panes = "B2"
    widths = {"email_id": 11, "subject_full": 40, "body_preview": 50, "proposed_category": 24,
              "proposal_method": 13, "v1_gold_category": 30, "gold_category": 32,
              "gold_inquiry_answer_source": 26, "gold_request_type": 28,
              "gold_booking_lifecycle_stage": 26, "gold_sender_type": 22,
              "gold_requires_human_followup": 26, "gold_urgency_signal": 20,
              "ambiguous": 10, "notes": 34}
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 16)
    wrap = Alignment(wrap_text=True, vertical="top")
    for col in ("subject_full", "body_preview", "v1_gold_category", "notes"):
        ci = COLUMNS.index(col) + 1
        for row in range(2, last + 1):
            ws.cell(row, ci).alignment = wrap
    # highlight the two fresh-label columns
    for col in ("gold_category", "gold_inquiry_answer_source"):
        ci = COLUMNS.index(col) + 1
        for row in range(2, last + 1):
            ws.cell(row, ci).fill = BLANK

    # hidden Lists sheet + data-validation dropdowns (request_type's 11 values exceed
    # Excel's 255-char inline-list limit, so reference ranges).
    lists = wb.create_sheet("Lists")
    lists.sheet_state = "hidden"
    for j, (name, values) in enumerate(FACETS, start=1):
        lcol = get_column_letter(j)
        lists.cell(1, j, name)
        for k, v in enumerate(values, start=2):
            lists.cell(k, j, v)
        dv = DataValidation(type="list", allow_blank=True, showErrorMessage=True,
                            formula1=f"Lists!${lcol}$2:${lcol}${1 + len(values)}")
        dv.errorTitle, dv.error = "Invalid label", "Pick a value from the dropdown."
        ws.add_data_validation(dv)
        gcol = get_column_letter(COLUMNS.index(name) + 1)
        dv.add(f"{gcol}2:{gcol}{last}")

    wb.save(V2_SHEET)
    print(f"v2 sheet: {V2_SHEET}  ({len(v1)} emails)")
    print("Fresh-label (highlighted): gold_category (10), gold_inquiry_answer_source (5)")
    print("Carried for review: request_type, lifecycle, requires_human_followup")
    print("Carried unchanged:  sender_type, urgency_signal | reference: v1_gold_category")
    print("Read bodies in: outputs/gold/gold_emails.md (same 49 emails)")


if __name__ == "__main__":
    main()