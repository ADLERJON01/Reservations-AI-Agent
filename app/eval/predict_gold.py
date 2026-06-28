"""Run the full pipeline over the gold-set emails and dump predictions for scoring.

LIVE — needs Ollama (it runs the real classifier). Small: 49 emails, not the 378
batch. Writes outputs/gold/gold_predictions.csv; the scorer (score_gold.py) is
offline and consumes that CSV, so you only pay the live cost once.

    python -m app.eval.predict_gold
"""
from __future__ import annotations

import csv
import warnings
from pathlib import Path

from openpyxl import load_workbook

from app.config import get_settings
from app.graph import run_pipeline

OUT_DIR = Path(__file__).resolve().parents[2] / "outputs" / "gold"
SHEET = OUT_DIR / "gold_labeling_sheet.xlsx"
RAW = get_settings().inputs_dir / "raw_emails"

PRED_COLS = [
    "email_id", "schema_valid", "pred_category", "pred_sender_type", "pred_request_type",
    "pred_inquiry_answer_source", "pred_booking_lifecycle_stage",
    "pred_requires_human_followup", "pred_urgency_signal",
    "recommended_action", "applied_rule_id", "model_name", "error",
]


def gold_ids() -> list[str]:
    warnings.filterwarnings("ignore")              # openpyxl drops the DV extension on read
    ws = load_workbook(SHEET)["labels"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    return [r[idx["email_id"]] for r in rows[1:] if r[idx["email_id"]]]


def predict_one(email_id: str) -> dict:
    state = run_pipeline(RAW / f"{email_id}.txt")
    cls = state.llm_output.classification if state.llm_output else None
    return {
        "email_id": email_id,
        "schema_valid": cls is not None,
        "pred_category": cls.predicted_category if cls else "",
        "pred_sender_type": cls.sender_type if cls else "",
        "pred_request_type": cls.request_type if cls else "",
        "pred_inquiry_answer_source": cls.inquiry_answer_source if cls else "",
        "pred_booking_lifecycle_stage": cls.booking_lifecycle_stage if cls else "",
        "pred_requires_human_followup": cls.requires_human_followup if cls else "",
        "pred_urgency_signal": cls.urgency_signal if cls else "",
        "recommended_action": state.recommended_action or "",
        "applied_rule_id": state.applied_rule_id or "",
        "model_name": state.model_name or "",
        "error": (state.errors[-1].get("message", "") if state.errors else ""),
    }


def main() -> None:
    ids = gold_ids()
    out = OUT_DIR / "gold_predictions.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=PRED_COLS)
        w.writeheader()
        for i, eid in enumerate(ids, 1):
            try:
                row = predict_one(eid)
            except Exception as e:                 # keep going; record the failure
                row = {c: "" for c in PRED_COLS}
                row.update(email_id=eid, schema_valid=False, error=str(e)[:200])
            w.writerow(row)
            print(f"[{i:2d}/{len(ids)}] {eid}: {row['pred_category'] or '(invalid)'}")
    print(f"\n→ {out}\nNow run:  python -m app.eval.score_gold")


if __name__ == "__main__":
    main()
