"""Score v2 predictions vs the v2 gold labels (OFFLINE — no Ollama).

Same per-facet / per-category reporting as score_gold.py, PLUS the headline v2 metric:
**RAG-candidate precision/recall** — does the AND-gate (knowledge_policy_inquiry AND
kb_policy) fire on the right emails? A FALSE RAG candidate (drafted from static KB when
the email actually needs live data/staff) is the dangerous error; report it explicitly.

Compares the category number against the v1 enriched baseline (79.6%). Writes
outputs/gold/gold_metrics_v2.md.

    python -m app.eval.score_gold_v2
"""
from __future__ import annotations

import csv
import warnings
from pathlib import Path

from openpyxl import load_workbook

OUT_DIR = Path(__file__).resolve().parents[2] / "outputs" / "gold"
SHEET = OUT_DIR / "gold_labeling_sheet_v2.xlsx"
PRED = OUT_DIR / "gold_predictions_v2.csv"

V1_CATEGORY_BASELINE = 0.796   # enriched v1 (outputs/gold/gold_metrics.md)

# facet label: (gold column, prediction column)
FACETS = [
    ("category", "gold_category", "pred_category"),
    ("request_type", "gold_request_type", "pred_request_type"),
    ("inquiry_answer_source", "gold_inquiry_answer_source", "pred_inquiry_answer_source"),
    ("lifecycle", "gold_booking_lifecycle_stage", "pred_booking_lifecycle_stage"),
    ("requires_human_followup", "gold_requires_human_followup", "pred_requires_human_followup"),
    ("sender_type", "gold_sender_type", "pred_sender_type"),
    ("urgency_signal", "gold_urgency_signal", "pred_urgency_signal"),
]


def _rag(cat: str, src: str) -> bool:
    return cat == "knowledge_policy_inquiry" and src == "kb_policy"


def load_gold() -> dict[str, dict]:
    warnings.filterwarnings("ignore")
    ws = load_workbook(SHEET)["labels"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    gold = {}
    for r in rows[1:]:
        eid = r[idx["email_id"]]
        if not eid:
            continue
        gold[eid] = {k: (r[idx[k]] or "") for k in idx}
        gold[eid]["_ambiguous"] = str(r[idx["ambiguous"]] or "").strip().upper() == "Y"
    return gold


def load_pred() -> dict[str, dict]:
    with open(PRED, newline="", encoding="utf-8") as f:
        return {row["email_id"]: row for row in csv.DictReader(f)}


def per_class(pairs: list[tuple[str, str]]) -> dict[str, tuple]:
    classes = sorted(set(g for g, _ in pairs) | set(p for _, p in pairs))
    out = {}
    for c in classes:
        tp = sum(1 for g, p in pairs if g == c and p == c)
        fp = sum(1 for g, p in pairs if g != c and p == c)
        fn = sum(1 for g, p in pairs if g == c and p != c)
        support = sum(1 for g, _ in pairs if g == c)
        prec = tp / (tp + fp) if tp + fp else 0.0
        rec = tp / (tp + fn) if tp + fn else 0.0
        f1 = 2 * prec * rec / (prec + rec) if prec + rec else 0.0
        out[c] = (prec, rec, f1, support)
    return out


def _gate_check(gold: dict) -> list[str]:
    """Sanity: did the labeler leave the fresh v2 fields blank anywhere?"""
    miss = [e for e, g in gold.items() if not g.get("gold_category")
            or not g.get("gold_inquiry_answer_source")]
    return miss


def main() -> None:
    if not PRED.exists():
        raise SystemExit(f"No predictions at {PRED}. Run: python -m app.eval.predict_gold_v2")
    gold, pred = load_gold(), load_pred()
    blank = _gate_check(gold)
    if blank:
        raise SystemExit(f"{len(blank)} emails still missing gold_category/gold_inquiry_answer_source "
                         f"— finish labeling first: {blank}")

    ids = [e for e in gold if e in pred]
    missing = [e for e in gold if e not in pred]
    lines: list[str] = ["# Gold-set metrics — v2.0.0 taxonomy\n",
                        f"Scored **{len(ids)}** emails"
                        + (f" ({len(missing)} missing predictions: {missing})" if missing else "")
                        + ".\n"]

    invalid = [e for e in ids if str(pred[e].get("schema_valid")).lower() != "true"]
    if invalid:
        lines.append(f"Schema-invalid predictions (count as wrong): {len(invalid)} — {invalid}\n")

    # --- per-facet accuracy ---
    lines.append("## Facet accuracy\n")
    lines.append("| facet | accuracy | correct/total |")
    lines.append("|---|---|---|")
    facet_acc = {}
    for name, gcol, pcol in FACETS:
        correct = sum(1 for e in ids if gold[e][gcol] == pred[e][pcol])
        acc = correct / len(ids) if ids else 0.0
        facet_acc[name] = acc
        lines.append(f"| {name} | {acc:.1%} | {correct}/{len(ids)} |")
    lines.append("")

    # --- category vs v1 ---
    cat_acc = facet_acc["category"]
    delta = cat_acc - V1_CATEGORY_BASELINE
    unamb = [e for e in ids if not gold[e]["_ambiguous"]]
    cat_acc_un = (sum(1 for e in unamb if gold[e]["gold_category"] == pred[e]["pred_category"])
                  / len(unamb)) if unamb else 0.0
    lines.append(f"**Category {cat_acc:.1%} (all) vs v1 enriched {V1_CATEGORY_BASELINE:.1%} "
                 f"→ {delta:+.1%}.** Unambiguous ({len(unamb)}): {cat_acc_un:.1%}. "
                 f"*(Note: v2 has 10 categories vs v1's 7 — a harder task; judge with RAG precision below.)*\n")

    # --- per-category P/R/F1 ---
    cat_pairs = [(gold[e]["gold_category"], pred[e]["pred_category"] or "(invalid)") for e in ids]
    pc = per_class(cat_pairs)
    lines.append("## Category precision / recall / F1\n")
    lines.append("| category | P | R | F1 | support |")
    lines.append("|---|---|---|---|---|")
    for c, (p, r, f1, s) in sorted(pc.items(), key=lambda kv: -kv[1][3]):
        lines.append(f"| {c} | {p:.2f} | {r:.2f} | {f1:.2f} | {s} |")
    macro = sum(v[2] for v in pc.values()) / len(pc) if pc else 0.0
    lines.append(f"\n**Macro-F1: {macro:.2f}**\n")

    # --- RAG-candidate precision/recall (the v2 safety metric) ---
    gold_rag = {e: _rag(gold[e]["gold_category"], gold[e]["gold_inquiry_answer_source"]) for e in ids}
    pred_rag = {e: _rag(pred[e]["pred_category"], pred[e]["pred_inquiry_answer_source"]) for e in ids}
    tp = [e for e in ids if gold_rag[e] and pred_rag[e]]
    fp = [e for e in ids if not gold_rag[e] and pred_rag[e]]      # DANGEROUS: false RAG candidate
    fn = [e for e in ids if gold_rag[e] and not pred_rag[e]]      # missed RAG candidate
    prec = len(tp) / (len(tp) + len(fp)) if (tp or fp) else 0.0
    rec = len(tp) / (len(tp) + len(fn)) if (tp or fn) else 0.0
    f1 = 2 * prec * rec / (prec + rec) if (prec + rec) else 0.0
    lines.append("## RAG-candidate gate (knowledge_policy_inquiry AND kb_policy)\n")
    lines.append(f"- gold RAG candidates: **{sum(gold_rag.values())}**  |  predicted: **{sum(pred_rag.values())}**")
    lines.append(f"- **precision {prec:.1%}** (TP {len(tp)} / TP+FP {len(tp)+len(fp)})  |  "
                 f"**recall {rec:.1%}** (TP {len(tp)} / TP+FN {len(tp)+len(fn)})  |  F1 {f1:.2f}")
    lines.append(f"- **FALSE RAG candidates (dangerous — would draft from KB) [{len(fp)}]:** {fp or 'none'}")
    lines.append(f"- missed RAG candidates (escalated instead) [{len(fn)}]: {fn or 'none'}\n")

    # --- category misclassifications ---
    errs = [(e, gold[e]["gold_category"], pred[e]["pred_category"] or "(invalid)")
            for e in ids if gold[e]["gold_category"] != (pred[e]["pred_category"] or "(invalid)")]
    lines.append(f"## Category misclassifications ({len(errs)})\n")
    for e, g, p in errs:
        flag = " [ambiguous]" if gold[e]["_ambiguous"] else ""
        lines.append(f"- `{e}`: gold **{g}** → pred **{p}**{flag}")
    lines.append("")

    out = OUT_DIR / "gold_metrics_v2.md"
    out.write_text("\n".join(lines), encoding="utf-8")

    # console summary
    print(f"scored {len(ids)} emails (v2)")
    print(f"category: {cat_acc:.1%} (all) | {cat_acc_un:.1%} (unamb) | macro-F1 {macro:.2f}"
          f"  [v1 enriched {V1_CATEGORY_BASELINE:.1%}, {delta:+.1%}]")
    print(f"RAG gate: precision {prec:.1%} | recall {rec:.1%} | "
          f"false-candidates {len(fp)} | missed {len(fn)}")
    for name in ("inquiry_answer_source", "request_type", "requires_human_followup",
                 "lifecycle", "sender_type", "urgency_signal"):
        print(f"  {name:24s} {facet_acc[name]:.1%}")
    print(f"→ {out}")


if __name__ == "__main__":
    main()