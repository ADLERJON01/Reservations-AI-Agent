"""Run the v1.1 classifier (v1 7-cat + inquiry_answer_source patch) over the gold emails.

LIVE — needs Ollama. Throwaway harness (no pipeline migration): calls the LLM client
directly with the v1.1 contract (EmailExtractionV11) + v1.1 prompt + preprocessor v2.
Writes outputs/gold/gold_predictions_v1_1.csv. Then: python -m app.eval.score_gold_v1_1

    python -m app.eval.predict_gold_v1_1
"""
from __future__ import annotations

import csv
import warnings
from pathlib import Path

from openpyxl import load_workbook

from app.agents.preprocessor import preprocess
from app.agents.prompts_v1_1 import build_system_prompt, build_user_prompt
from app.config import get_settings
from app.llm.ollama_native import OllamaNativeClient
from app.models.llm_output_v1_1 import EmailExtractionV11

OUT_DIR = Path(__file__).resolve().parents[2] / "outputs" / "gold"
SHEET = OUT_DIR / "gold_labeling_sheet_v2.xlsx"          # same 49 emails; category collapsed at scoring
RAW = get_settings().inputs_dir / "raw_emails"

PRED_COLS = [
    "email_id", "schema_valid", "pred_category", "pred_request_type",
    "pred_inquiry_answer_source", "pred_sender_type", "pred_booking_lifecycle_stage",
    "pred_requires_human_followup", "pred_urgency_signal", "pred_rag_candidate",
    "model_name", "error",
]


def _rag_candidate(cls) -> bool:
    """Derived in code (v1.1 AND-gate)."""
    return cls.category == "service_or_information_inquiry" and cls.inquiry_answer_source == "kb_policy"


def gold_ids() -> list[str]:
    warnings.filterwarnings("ignore")
    ws = load_workbook(SHEET)["labels"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    return [r[idx["email_id"]] for r in rows[1:] if r[idx["email_id"]]]


def classify_v1_1(email, client, s):
    system_prompt, user_prompt = build_system_prompt(), build_user_prompt(email)
    last_error = None
    for attempt in range(s.classifier_max_retries + 1):
        temperature = s.classifier_temperature if attempt == 0 else s.classifier_retry_temperature
        seed = s.classifier_seed + attempt
        r = client.call_structured(system_prompt, user_prompt, response_model=EmailExtractionV11,
                                   model=s.primary_model, temperature=temperature, seed=seed)
        if r.valid and r.output is not None:
            return r.output, None
        last_error = r.error
    return None, last_error


def predict_one(email_id, client, s) -> dict:
    email = preprocess(RAW / f"{email_id}.txt")
    out, err = classify_v1_1(email, client, s)
    cls = out.classification if out else None
    return {
        "email_id": email_id,
        "schema_valid": cls is not None,
        "pred_category": cls.category if cls else "",
        "pred_request_type": cls.request_type if cls else "",
        "pred_inquiry_answer_source": cls.inquiry_answer_source if cls else "",
        "pred_sender_type": cls.sender_type if cls else "",
        "pred_booking_lifecycle_stage": cls.booking_lifecycle_stage if cls else "",
        "pred_requires_human_followup": cls.requires_human_followup if cls else "",
        "pred_urgency_signal": cls.urgency_signal if cls else "",
        "pred_rag_candidate": _rag_candidate(cls) if cls else False,
        "model_name": s.primary_model,
        "error": (err or "")[:200],
    }


def main() -> None:
    s = get_settings()
    client = OllamaNativeClient()
    ids = gold_ids()
    out = OUT_DIR / "gold_predictions_v1_1.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=PRED_COLS)
        w.writeheader()
        for i, eid in enumerate(ids, 1):
            try:
                row = predict_one(eid, client, s)
            except Exception as e:
                row = {c: "" for c in PRED_COLS}
                row.update(email_id=eid, schema_valid=False, error=str(e)[:200])
            w.writerow(row)
            tag = "  <RAG>" if row.get("pred_rag_candidate") else ""
            print(f"[{i:2d}/{len(ids)}] {eid}: {row['pred_category'] or '(invalid)'}{tag}")
    print(f"\n→ {out}\nNow run:  python -m app.eval.score_gold_v1_1")


if __name__ == "__main__":
    main()