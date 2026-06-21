"""Score pipeline predictions against the gold labels (OFFLINE — no Ollama).

Reads the labeled gold_labeling_sheet.xlsx + gold_predictions.csv (produced by
predict_gold.py), and reports: category accuracy (overall + unambiguous subset),
per-category precision/recall/F1 + macro-F1, the misclassifications, and per-facet
accuracy. Writes outputs/gold/gold_metrics.md.

    python -m app.eval.score_gold
"""
from __future__ import annotations

import csv
import warnings
from pathlib import Path

from openpyxl import load_workbook

OUT_DIR = Path(__file__).resolve().parents[2] / "outputs" / "gold"
SHEET = OUT_DIR / "gold_labeling_sheet.xlsx"
PRED = OUT_DIR / "gold_predictions.csv"

# facet label: (gold column, prediction column)
FACETS = [
    ("category", "gold_category", "pred_category"),
    ("sender_type", "gold_sender_type", "pred_sender_type"),
    ("request_type", "gold_request_type", "pred_request_type"),
    ("lifecycle", "gold_booking_lifecycle_stage", "pred_booking_lifecycle_stage"),
    ("expects_human_response", "gold_expects_human_response", "pred_expects_human_response"),
    ("urgency_signal", "gold_urgency_signal", "pred_urgency_signal"),
]


def load_gold() -> dict[str, dict]:
    warnings.filterwarnings("ignore")
    ws = load_workbook(SHEET)["labels"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    gold = {}
    for r in rows[1:]:
        eid = r[idx["email_id"]]
        if not eid:
            continue
        gold[eid] = {k: (r[idx[k]] or "") for k in idx}
        gold[eid]["_ambiguous"] = str(r[idx["ambiguous"]] or "").strip().upper() == "Y"
    return gold


def load_pred() -> dict[str, dict]:
    with open(PRED, newline="", encoding="utf-8") as f:
        return {row["email_id"]: row for row in csv.DictReader(f)}


def per_class(pairs: list[tuple[str, str]]) -> dict[str, tuple]:
    """precision, recall, f1, support per gold class."""
    classes = sorted(set(g for g, _ in pairs) | set(p for _, p in pairs))
    out = {}
    for c in classes:
        tp = sum(1 for g, p in pairs if g == c and p == c)
        fp = sum(1 for g, p in pairs if g != c and p == c)
        fn = sum(1 for g, p in pairs if g == c and p != c)
        support = sum(1 for g, _ in pairs if g == c)
        prec = tp / (tp + fp) if tp + fp else 0.0
        rec = tp / (tp + fn) if tp + fn else 0.0
        f1 = 2 * prec * rec / (prec + rec) if prec + rec else 0.0
        out[c] = (prec, rec, f1, support)
    return out


def main() -> None:
    if not PRED.exists():
        raise SystemExit(f"No predictions at {PRED}. Run: python -m app.eval.predict_gold")
    gold, pred = load_gold(), load_pred()
    ids = [e for e in gold if e in pred]
    missing = [e for e in gold if e not in pred]

    lines: list[str] = ["# Gold-set metrics\n",
                        f"Scored **{len(ids)}** emails"
                        + (f" ({len(missing)} missing predictions: {missing})" if missing else "")
                        + ".\n"]

    invalid = [e for e in ids if str(pred[e].get("schema_valid")).lower() != "true"]
    if invalid:
        lines.append(f"Schema-invalid predictions (count as wrong): {len(invalid)} — {invalid}\n")

    # --- per-facet accuracy ---
    lines.append("## Facet accuracy\n")
    lines.append("| facet | accuracy | correct/total |")
    lines.append("|---|---|---|")
    for name, gcol, pcol in FACETS:
        correct = sum(1 for e in ids if gold[e][gcol] == pred[e][pcol])
        acc = correct / len(ids) if ids else 0.0
        lines.append(f"| {name} | {acc:.1%} | {correct}/{len(ids)} |")
    lines.append("")

    # --- category: unambiguous subset ---
    unamb = [e for e in ids if not gold[e]["_ambiguous"]]
    cat_acc_un = (sum(1 for e in unamb if gold[e]["gold_category"] == pred[e]["pred_category"])
                  / len(unamb)) if unamb else 0.0
    lines.append(f"**Category accuracy on the {len(unamb)} unambiguous emails: "
                 f"{cat_acc_un:.1%}** (excludes the {len(ids) - len(unamb)} flagged `ambiguous`).\n")

    # --- per-category precision/recall/f1 ---
    cat_pairs = [(gold[e]["gold_category"], pred[e]["pred_category"] or "(invalid)") for e in ids]
    pc = per_class(cat_pairs)
    lines.append("## Category precision / recall / F1\n")
    lines.append("| category | P | R | F1 | support |")
    lines.append("|---|---|---|---|---|")
    for c, (p, r, f1, s) in sorted(pc.items(), key=lambda kv: -kv[1][3]):
        lines.append(f"| {c} | {p:.2f} | {r:.2f} | {f1:.2f} | {s} |")
    macro = sum(v[2] for v in pc.values()) / len(pc) if pc else 0.0
    lines.append(f"\n**Macro-F1: {macro:.2f}**\n")

    # --- misclassifications ---
    errs = [(e, gold[e]["gold_category"], pred[e]["pred_category"] or "(invalid)")
            for e in ids if gold[e]["gold_category"] != (pred[e]["pred_category"] or "(invalid)")]
    lines.append(f"## Category misclassifications ({len(errs)})\n")
    for e, g, p in errs:
        flag = " [ambiguous]" if gold[e]["_ambiguous"] else ""
        lines.append(f"- `{e}`: gold **{g}** → pred **{p}**{flag}")
    lines.append("")

    out = OUT_DIR / "gold_metrics.md"
    out.write_text("\n".join(lines), encoding="utf-8")

    # console summary
    cat_acc = sum(1 for g, p in cat_pairs if g == p) / len(cat_pairs) if cat_pairs else 0.0
    print(f"scored {len(ids)} emails")
    print(f"category accuracy: {cat_acc:.1%} (all)  |  {cat_acc_un:.1%} (unambiguous)  |  macro-F1 {macro:.2f}")
    for name, gcol, pcol in FACETS[1:]:
        acc = sum(1 for e in ids if gold[e][gcol] == pred[e][pcol]) / len(ids)
        print(f"  {name:24s} {acc:.1%}")
    print(f"→ {out}")


if __name__ == "__main__":
    main()
